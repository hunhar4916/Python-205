studentID = "hunhar4916"
print(studentID)
print("Its just a flesh wound!")

"""
1. Ask user for 5 numbers
2. Ask user for 5 names/incomes and appends them to final.csv
3.Builds a pie chart in Excel
4. Builds a vertical bar graph
"""

import csv
import datetime
import openpyxl
from openpyxl.chart import PieChart, Reference
import matplotlib.pyplot as plt

#Path to existing data file
FILE_PATH = r"C:\FinalExam\final.csv"
EXCEL_PATH = r"C:\FinalExam\final.xlsx"

def askUser():
    #Asks user for 5 nums, gets their total, and displays the result
    total = 0

    for i in range(0, 5):
        number = float(input("Please enter a number: "))
        total += number

    print(f"The sum of the 5 numbers entered is: {total}")
    return total

def askIncome():
    #Asks the user for 5 names and incomes and appends each to final.csv
    with open(FILE_PATH, "a", newline="") as file:
        for i in range(0, 5):
            name = input("Please enter a name: ")
            income = input("Please enter their income: ")
            file.write(f"\n{name},{income}")


def excelPie():
    #Reads data from final.csv and builds a pie chart for it
    workbook = openpyxl.Workbook()

    sheet = workbook.active

    with open(FILE_PATH, "r") as file:
        reader = csv.reader(file)
        for row in reader:
            if row:
                #Cast the income value into a number
                name, income = row[0], int(row[1])
                #write the name/income pair into the next row
                sheet.append([name, income])

    today = datetime.date.today().strftime("%B %d, %Y")
    chart_title = f"{studentID} {today}"

    pie = PieChart()
    pie.title = chart_title

    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    labels = Reference(sheet, min_col=1, min_row=1, max_row=sheet.max_row)

    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)

    #Place the finshed pie chart
    sheet.add_chart(pie, "D2")
    workbook.save(EXCEL_PATH)

def verticalBar():
    #Reads all data from final.csv and builds a bar graph
    names = []
    income = []


    with open(FILE_PATH, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            if row:
                names.append(row[0])
                income.append(row[1])

    today = datetime.date.today().strftime("%B %d, %Y")
    chart_title = f"{studentID} {today}"

    plt.bar(names, income, color="pink", label="Incomes")
    plt.title(chart_title)
    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.legend()
    plt.show()


if __name__ == "__main__":
    askUser()
    askIncome()
    excelPie()
    verticalBar()
    














    


































    













                
                
















    
        






















