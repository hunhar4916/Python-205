from datetime import datetime
import csv

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

print("hunhar4916 - Spreadsheet Automation Menu")
print("It's just a flesh wound!!")

STUDENT_ID = "hunhar4916"


def menu_options():
    """
    menu_options(): prints the main menu options available to the user.
    """
    options = [
        "1. input data",
        "2. Current data",
        "3. Generate report",
        "0. Exit program"
    ]
    print("\n--Menu--")
    for option in options:
        print(option)


def convertData(data):
    """
    convertData(data): takes one numerical argument (data), the
    temperature in Fahrenheit, and returns it converted to Celsius
    using (F - 32) * 5/9. Returns a float.
    """
    return (data - 32) * 5 / 9


def insertData(path, data):
    """
    insertData(path, data): takes a path to a csv file and a comma-separated
    string of data. Opens the file with append permissions (creating the
    file first if it does not already exist) and writes the data as a new
    line. A try-except statement catches any errors that occur while
    writing; if an error occurs, it is reported and re-raised so the
    caller can respond accordingly.
    """
    try:
        with open(path, "a") as file:
            file.write(data + "\n")
    except Exception as e:
        print(f"Error writing to file {path}: {e}")
        raise


def viewData(path):
    """
    viewData(path): takes a path to a csv file, opens it using the minimal
    permissions needed to read (read-only mode), and displays the path of
    the file being read along with its full contents. A try-except
    statement catches any errors that occur while reading, such as the
    file not existing.
    """
    try:
        with open(path, "r") as file:
            print(f"\nReading data from: {path}")
            print(file.read())
    except Exception as e:
        print(f"Error reading file {path}: {e}")


def getInput():
    """
    getInput(): prompts the user for the number of entries to input, then
    for each entry collects a date and a highest Fahrenheit temperature,
    converts the temperature to Celsius via convertData, and attempts to
    save the entry to ZooData.csv using insertData. A try-except statement
    wraps the save so that, on a successful write, a confirmation message
    is displayed showing the current date/time and the data saved; on
    failure, an error message is displayed instead.
    """
    num_entries = int(input("How many entries are you inputting? "))
    for i in range(num_entries):
        entry_date = input("\nEnter a date: ")
        print()
        highest_temp = float(input("Enter the highest temp for the inputted date: "))
        # Calling convertData(highest_temp): passes the Fahrenheit value
        # to be converted; returns the converted Celsius value (float)
        converted_value = convertData(highest_temp)
        data = f"{entry_date},{highest_temp},{converted_value}"

        try:
            insertData("ZooData.csv", data)
            print(f"The following data was saved at {datetime.now()}: {data}.")
        except Exception as e:
            print(f"Failed to save entry: {e}")


def createChart(path, chart_type):
    """
    createChart(path, chart_type): takes two arguments - path, a string
    containing the path to the csv data file, and chart_type, a string
    that is either "line" or "bar" indicating which kind of chart to
    build. Asks the user whether to chart the initial data (Fahrenheit)
    or the converted data (Celsius), reads the dates and the selected
    values out of the csv file (casting the temperature values to
    float), writes that data into a new Excel workbook named final.xlsx
    using openpyxl, and inserts a bar or line chart into the workbook
    using the values as the data series and the dates as the category
    labels. The chart's axes are labeled and its title is set to
    "<student ID> <current date>". The completed workbook is saved to
    final.xlsx. This function does not return a value.
    """
    source_choice = input(
        "\nWhich data source would you like to chart?\n"
        "1. Initial data (Fahrenheit)\n"
        "2. Converted data (Celsius)\n"
        "Enter your choice (1-2): "
    )

    if source_choice == "1":
        value_index = 1
        value_label = "Temperature (Fahrenheit)"
    elif source_choice == "2":
        value_index = 2
        value_label = "Temperature (Celsius)"
    else:
        print("Invalid selection. Defaulting to initial data (Fahrenheit).")
        value_index = 1
        value_label = "Temperature (Fahrenheit)"

    dates = []
    values = []

    try:
        with open(path, "r") as file:
            reader = csv.reader(file)
            for row in reader:
                if not row:
                    continue
                dates.append(row[0])
                values.append(float(row[value_index]))
    except Exception as e:
        print(f"Error reading file {path}: {e}")
        return

    # Build the workbook and write the raw data to a worksheet.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Date", value_label])
    for entry_date, value in zip(dates, values):
        sheet.append([entry_date, value])

    chart_title = f"{STUDENT_ID} {datetime.now().strftime('%m/%d/%Y')}"

    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "line":
        chart = LineChart()
    else:
        print(f"Unsupported chart type: {chart_type}")
        return

    chart.title = chart_title
    chart.x_axis.title = "Date"
    chart.y_axis.title = value_label

    data_ref = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    sheet.add_chart(chart, "D2")

    try:
        workbook.save("final.xlsx")
        print(f"\nChart saved to final.xlsx as a {chart_type} chart titled '{chart_title}'.")
    except Exception as e:
        print(f"Error saving final.xlsx: {e}")


def generateReport(path):
    """
    generateReport(path): takes one argument, path, a string containing
    the path to the csv data file. Asks the user which chart type they
    would like to generate (line or bar) and calls createChart, passing
    along the csv path and the chosen chart type. This function does not
    return a value.
    """
    chart_choice = input(
        "\nWhich graph type would you like to create?\n"
        "1. Line chart\n"
        "2. Bar chart\n"
        "Enter your choice (1-2): "
    )

    if chart_choice == "1":
        createChart(path, "line")
    elif chart_choice == "2":
        createChart(path, "bar")
    else:
        print("Invalid selection. Please choose 1 or 2.")


def current_data():
    """
    current_data(): placeholder for viewing current data (superseded by
    the viewData function, which is now called directly from the menu).
    """
    print("viewing current data... (feature coming soon)")


def exit_program():
    """
    exit_program(): prints a farewell message when the user exits the
    program.
    """
    print("Goodbye!")


while True:
    menu_options()
    # The next line retrieves the inputted option and stores into the
    # variable called choice.
    choice = input("Enter your choice (1-3): ")
    if choice == "1":
        print(f"You selected {choice} at {datetime.now()}")
        getInput()
    elif choice == "2":
        print(f"You selected {choice} at {datetime.now()}")
        viewData("ZooData.csv")
    elif choice == "3":
        print(f"You selected {choice} at {datetime.now()}")
        generateReport("ZooData.csv")
    elif choice == "0":
        exit_program()
        break
    else:
        print("Error: The chosen functionality is not implemented yet")
